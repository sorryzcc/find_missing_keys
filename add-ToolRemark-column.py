import openpyxl
from openpyxl.styles import Alignment

# 输入和输出文件路径
input_file_path = 'opsXY2.xlsx'  # 原始表格文件
output_file_path = 'opsXY2_with_remarks.xlsx'  # 更新后的表格文件

# 打开 Excel 文件
workbook = openpyxl.load_workbook(input_file_path)
sheet = workbook.active  # 默认处理第一个工作表

# 获取表头
headers = [cell.value for cell in sheet[1]]
headers.append('备注')  # 新增“备注”列

# 在第一行插入新表头
sheet.insert_cols(len(headers))  # 插入新列
sheet.cell(row=1, column=len(headers)).value = '备注'

# 处理每一行数据
for row_idx in range(2, sheet.max_row + 1):  # 从第二行开始（跳过表头）
    tool_remark = sheet.cell(row=row_idx, column=headers.index('ToolRemark') + 1).value or ''
    version = sheet.cell(row=row_idx, column=headers.index('Version') + 1).value or ''
    po = sheet.cell(row=row_idx, column=headers.index('PO') + 1).value or ''

    # 根据逻辑生成“备注”列，并添加换行符 \n
    remark = f"场景：{tool_remark}\n使用版本：{version}\n负责人：{po}"
    sheet.cell(row=row_idx, column=len(headers)).value = remark

    # 设置单元格样式以支持换行
    sheet.cell(row=row_idx, column=len(headers)).alignment = Alignment(wrap_text=True)

# 调整列宽以便更好地显示内容
for col in range(1, len(headers) + 1):
    sheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20

# 保存更新后的 Excel 文件
workbook.save(output_file_path)

print(f"新文件已生成：{output_file_path}")